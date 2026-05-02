"""Read Glow Visage Apr + May xlsx → daily records.

Re-reads files from disk every call (caller decides caching strategy).
"""
import os
import re
from pathlib import Path
from typing import Tuple

import openpyxl

APR_PATH = Path("/Users/fadiion/Documents/Claude/Projects/Ads & Report/Action Plan Glow - Apr 2026.xlsx")
MAY_PATH = Path("/Users/fadiion/Documents/Claude/Projects/Ads & Report/Action Plan Glow - May 2026.xlsx")


def file_signature() -> Tuple[float, float]:
    """Return (apr_mtime, may_mtime) so callers can use as cache key."""
    apr = APR_PATH.stat().st_mtime if APR_PATH.exists() else 0
    may = MAY_PATH.stat().st_mtime if MAY_PATH.exists() else 0
    return (apr, may)


def _row_dict(ws, r, columns) -> dict:
    return {label: ws.cell(r, col).value for label, col in columns.items()}


def _parse_april():
    if not APR_PATH.exists():
        return []
    wb = openpyxl.load_workbook(APR_PATH, data_only=True)
    ws = wb["April"]

    day_starts = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str):
            m = re.match(r"Ads of\s+Day\s+(\d+)", v.strip())
            if m:
                day_starts.append((int(m.group(1)), r))

    cols = {
        "campaign": 3, "status": 2, "objective": 5, "budget": 8,
        "impression": 10, "reach": 14, "cpm": 17, "conversion": 19,
        "roas": 21, "result": 23, "cost_per_result": 27, "spent": 29,
    }

    out = []
    for i, (day, start_r) in enumerate(day_starts):
        ad_start = start_r + 6
        ad_end = day_starts[i + 1][1] - 1 if i + 1 < len(day_starts) else start_r + 12
        ads = []
        for r in range(ad_start, ad_end + 1):
            no = ws.cell(r, 1).value
            camp = ws.cell(r, 3).value
            if no is None and not camp:
                continue
            row = _row_dict(ws, r, cols)
            row["campaign"] = str(row["campaign"] or "").strip()
            for k in ("budget", "impression", "reach", "cpm", "conversion", "roas",
                      "result", "cost_per_result", "spent"):
                row[k] = row[k] or 0
            ads.append(row)
        if ads:
            out.append({"date": f"2026-04-{day:02d}", "month": "Apr", "day": day, "ads": ads})
    return out


def _parse_may():
    if not MAY_PATH.exists():
        return []
    wb = openpyxl.load_workbook(MAY_PATH, data_only=True)
    if "📅 Daily Log" not in wb.sheetnames:
        return []
    ws = wb["📅 Daily Log"]

    cols = {
        "status": 2, "objective": 3, "campaign": 4, "budget": 6,
        "impression": 7, "reach": 8, "cpm": 9, "conversion": 10,
        "roas": 11, "result": 12, "cost_per_result": 13, "spent": 14,
    }

    out = []
    for d in range(1, 32):
        strip = 5 + (d - 1) * 5
        if strip + 4 > ws.max_row:
            break
        ads = []
        for ar in range(strip + 1, strip + 4):
            no = ws.cell(ar, 1).value
            camp = ws.cell(ar, 4).value
            if no is None and not camp:
                continue
            row = _row_dict(ws, ar, cols)
            row["campaign"] = str(row["campaign"] or "").strip()
            for k in ("budget", "impression", "reach", "cpm", "conversion", "roas",
                      "result", "cost_per_result", "spent"):
                row[k] = row[k] or 0
            ads.append(row)
        if ads:
            out.append({"date": f"2026-05-{d:02d}", "month": "May", "day": d, "ads": ads})
    return out


def load_history() -> list:
    """Read Apr + May xlsx fresh from disk."""
    return _parse_april() + _parse_may()
